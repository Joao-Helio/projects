import requests
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta

# Set Habitica API credentials and path to save spreadsheet
user_id = "YOUR_USER_ID"
api_token = "YOUR_API_TOKEN"
file_path = "YOUR_FILE_PATH"

# Define API endpoints
base_url = "https://habitica.com/api/v3"
headers = {
    "x-api-user": user_id,
    "x-api-key": api_token,
}

# function to convert dates from Unix TimeStamp to ISO8601
def convert_unix_timestamp_to_date(timestamp):
    return datetime.utcfromtimestamp(timestamp / 1000).strftime('%Y-%m-%d')

# Create Workbook/Sheet
os.chdir(file_path)
workbook = Workbook()
sheet = workbook.active

# variable to paint the cells to represent checked habits for each day
greenFill = PatternFill(start_color='00008000',
                   end_color='00008000',
                   fill_type='solid')

# create a list of the last 30 days
last_30_days = []
for day in range(30):
    current_day = datetime.today() - timedelta(days=day)
    current_day = str(current_day)[0:10]
    last_30_days.insert(0, current_day)

# add spreadsheet headings
sheet.cell(row=1, column=1).value = 'Habits'
for i in range(1, 31):
    day = 1
    sheet.cell(row = 1, column = i+1).value = last_30_days[i-1]

# function get the list of dates on which a habit has been checked
def get_habit_dates(habit_id):
    habit_url = f"{base_url}/tasks/{habit_id}"

    try:
        response = requests.get(habit_url, headers=headers)
        response.raise_for_status()
        data = response.json()
        history = data["data"]["history"]
        checked_dates = [convert_unix_timestamp_to_date(entry["date"]) for entry in history if entry["value"] != 0]
        return checked_dates
    except requests.exceptions.RequestException as e:
        print(f"Error: {e}")
        return None
    
# function to return to habit name from a given ID
def get_habit_name(habit_id):
    habit_url = f"{base_url}/tasks/{habit_id}"
    try:
        response = requests.get(habit_url, headers=headers)
        response.raise_for_status()
        data = response.json()
        habit_name = data["data"]["text"]
        return habit_name
    except requests.exceptions.RequestException as e:
        print(f"Error: {e}")
        return None

# function to create a list of unique habits
def habits_list():
    response = requests.get(f"{base_url}/tasks/user", headers=headers)
    if response.status_code == 200:
        tasks = response.json()["data"]
        habits = [task for task in tasks]
        unique_habits = set()
        for habit in habits:
            ide = habit['id']
            type = habit['type']
            if type == 'habit' or type == 'daily':
                unique_habits.add(ide)
        return list(unique_habits)
    else:
        print("Error:", response.text)
        return []

# fill the spreadsheet first column and color the correct cells
def habits_history(habits_list):
    try:
        for i in range(len(habits_list)):
            checked_dates = get_habit_dates(habits_list[i])
            sheet.cell(row = i+2, column = 1).value = get_habit_name(habits_list[i])
            for j in range(2,32):
                if sheet.cell(row = 1, column = j).value in checked_dates:
                    sheet.cell(row = i+2, column = j).fill = greenFill
    except requests.exceptions.RequestException as e:
        print(f"Error: {e}")
        return None

habits_list = habits_list()
habits_history(habits_list)

# Formatting: Iterate over all columns and adjust their widths
for column in sheet.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    sheet.column_dimensions[column_letter].width = max_length

# save file as a new spreadsheet each day
save_path = file_path + '\\HabitsRecord - ' + str(datetime.today())[0:10] + '.xlsx'
workbook.save(save_path)